"""
Version: 1.0

Summary: compute the graph from simplified cross section image sequence 

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE:

python3 graph_compute.py -p /home/suxingliu/ply_data/CT_surface/ -ft png


argument:
("-p", "--path", required = True,    help = "path to image file")
("-ft", "--filetype", required = False, default = 'png',   help = "Image filetype")

"""

import glob, os, sys
import argparse

import numpy as np

from skimage.morphology import skeletonize
import sknw
import matplotlib.pyplot as plt
from mayavi import mlab
from mpl_toolkits.mplot3d import axes3d, Axes3D 
import cv2


import imagesize 
import progressbar
from time import sleep


from skimage.morphology import skeletonize_3d
from network_3d import skel2graph, plot_graph
from networkx import nx

import math

from openpyxl import load_workbook
from openpyxl import Workbook


def mkdir(path):
    """Create result folder"""
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False



#load image sequence
def load_image(image_file):
    
    path, filename = os.path.split(image_file)

    #base_name = os.path.splitext(os.path.basename(filename))[0]

    im_gray = cv2.imread(image_file, cv2.IMREAD_GRAYSCALE)

    #Obtain the threshold image using OTSU adaptive filter
    thresh = cv2.threshold(im_gray, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]

    return (thresh/255)
        
#get the color mapping
def get_cmap(n, name = 'hsv'):
    
    #Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    #RGB color; the keyword argument name must be a standard mpl colormap name
    
    return plt.cm.get_cmap(name, n)

#visualize graph using mayavi
def mayavi_visualize(graph,image_chunk):
    
    mlab.figure(1, size = (500, 500))
    
    cmap = get_cmap(len(graph.edges()))
    
    #for (s,e) in graph.edges():
        
    for idx, (s,e) in enumerate(graph.edges()):
        
         #generate different colors for different traces
        color_rgb = cmap(idx)[:len(cmap(idx))-1]
        
        pst = graph[s][e]['pts']

        mlab.plot3d(pst[:,0], pst[:,1], pst[:,2], tube_radius = 0.025, color = color_rgb)
    
    nodes = graph.nodes()
    
    ps = np.array([nodes[i]['o'] for i in nodes])
    
    mlab.points3d(ps[:,0], ps[:,1], ps[:,2], color=(1.0, 0.0, 0.0), scale_factor = 2.75)
    
    mlab.contour3d(image_chunk.astype(np.float),contours=[.5],opacity=.5,color=(1,1,1))
    
    mlab.show() 


#visualize graph using matplotlib
def plt_visualize(graph):
    
    fig = plt.figure()
    
    ax = fig.add_subplot(111, projection = '3d')
    
    for (s,e) in graph.edges():
        
        pst = graph[s][e]['pts']
        
        ax.plot(pst[:,0], pst[:,1], pst[:,2], c = 'green')
        
    nodes = graph.nodes()
    
    ps = np.array([nodes[i]['o'] for i in nodes])

    ax.scatter(ps[:,0], ps[:,1], ps[:,2], c = 'r')

    plt.title('Build Graph')
    
    plt.show()


def asSpherical(x, y, z):
    """coordinates transormation from cartesian coords to sphere coord system"""

    r = math.sqrt(x*x + y*y + z*z)
    
    elevation = math.acos(z/r)*180/math.pi #to degrees
    
    azimuth = np.arctan2(y,x)*180/math.pi
    
    return r, elevation, azimuth

def trace_angle(x, y, z):
    """compute the angle of each trace in 3D space"""
    
    #print(x[0],y[0],z[0])
    #print(x[len(x)-1],y[len(y)-1],z[len(z)-1])
    
    cx = x[0] - x[len(x)-1]
    cy = y[0] - y[len(y)-1]
    cz = z[0] - z[len(z)-1]

    (r,theta,phi) = asSpherical(cx, cy, cz)
    
    return r, theta, phi
    

if __name__ == '__main__':
    
    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True,    help = "path to image file")
    ap.add_argument("-ft", "--filetype", required = False, default = 'png',   help = "Image filetype")
    args = vars(ap.parse_args())

    #global file_path, save_path_ac, save_path_label, parent_path, pattern_id, count, whorl_dis_array, save_path_convex
    
    # setting path to cross section image files
    file_path = args["path"]
    ext = args['filetype']
     
    #accquire image file list
    filetype = '*.' + ext
    image_file_path = file_path + filetype

    #accquire image file list
    imgList = sorted(glob.glob(image_file_path))
    
    #print(imgList)
    
     # make the folder to store the results
    parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    mkpath = parent_path + '/' + str('analysis_result')
    mkdir(mkpath)
    save_path_result = mkpath + '/'
    
    
    n_samples = len(imgList)
    
    if n_samples > 0 :
        
        width, height = imagesize.get(imgList[0])
        #print(width, height)
    
    else:
        
        print("Empty image folder, abort!")
        sys.exit(0)
    
   
    #bar = progressbar.ProgressBar(maxval = n_samples, widgets = [progressbar.Bar('=', '[', ']'), ' ', progressbar.Percentage()])
    
    #progress bar display
    bar = progressbar.ProgressBar(maxval = n_samples)
    
    #bar.start()
    
    print("Loading images...")
   
    #initialize empty numpy array
    #image_chunk = np.empty(shape = (n_samples, height, width), dtype = np.float64)
    image_chunk = np.empty(shape = (n_samples, height, width), dtype = np.float64)
    
    #fill 3d image chunk with binary image data
    for file_idx, image_file in enumerate(imgList):
        
        image_chunk[file_idx, :, :] = load_image(image_file)
        
        bar.update(file_idx+1)
        
        sleep(0.1)

    bar.finish()
    
    print("image chunk size : {0}".format(str(image_chunk.shape)))
    
    #count number and frequency of 1 and 0
    uniqueValues, occurCount = np.unique(image_chunk, return_counts=True)
        
    print("Unique Values : " , uniqueValues)
        
    print("Occurrence Count : ", occurCount)
    
    
    #skeletonize
    skel = skeletonize(image_chunk)
    
    '''
    #build graph from skeleton
    graph = sknw.build_sknw(skel)
    
    mayavi_visualize(graph, image_chunk)
    '''
    
    
    skel = skel.astype(np.bool) #data needs to be bool
    
    G = skel2graph(skel) #create graph
    
    #print(G.size())
    
    
    #compute edge properities 
    numer_total = G.size()
    
    #plot the graph, use the z component to colorcode both the edges and the nodes, scale nodes according to their degree
    #plot_graph(G,node_color_keyword='z',edge_color_keyword='z',scale_node_keyword='degree')
    
    tube_surf, pts, edge_node_n1_select, edge_node_n2_select = plot_graph(G, node_color_keyword='z', edge_color_keyword='z')
    
    #print("edge_node_unique: {0}\n".format(edge_node_unique))
    
    edge_length = []
    edge_angle = []
    edgecount = 0
    index = []
    projection_radius = []
    
    
    for n1, n2, edgedict in G.edges(data = True):
        

        if (edge_node_n1_select[edgecount] == n1) and (edge_node_n2_select[edgecount] == n2) :
            
            edgecount += 1
            
            print("Properities of root index {0}:".format(edgecount))
            
            print("node1 = {0}, node2 = {1}, length = {2} ".format(n1, n2, edgedict['weight']))
            
            (r_data, theta_data, phi_data) = trace_angle(edgedict['x'], edgedict['y'], edgedict['z'])
            
            #test
            #x = [1,0]
            #(r_data, theta_data, phi_data) = trace_angle(x, x, x)
            
            print("r_data = {0}, theta_data = {1}, phi_data = {2}:\n".format(r_data, theta_data, phi_data))
            
            index.append(edgecount)
            edge_length.append(edgedict['weight'])
            edge_angle.append(phi_data)
            projection_radius.append(r_data)
    
                
    #show the binary data
    #mlab.contour3d(image_chunk.astype(np.float),contours=[.5],opacity=.5,color=(1,1,1))
    
    mlab.show()
    
    

    
    print ("results_folder: " + save_path_result) 
    ##################################################################
    #Start of writing measured parameters as excel file 
    
    trait_file = (save_path_result + 'root_trace_measure' + '.xlsx')
    
    trait_file_csv = (save_path_result + 'root_trace_measure' + '.csv')
    
    if os.path.exists(trait_file):
        # update values
        #Open an xlsx for reading
        wb = load_workbook(trait_file, read_only = False)
        sheet = wb.active

        os.remove(trait_file)
        
    else:
        # Keep presents
        wb = Workbook()
        sheet = wb.active
        
        sheet.cell(row = 1, column = 1).value = 'Root trace index'
        sheet.cell(row = 1, column = 2).value = 'Root trace length'
        sheet.cell(row = 1, column = 3).value = 'Root trace angle'
        #sheet.cell(row = 1, column = 4).value = 'Root trace diameter'
        sheet.cell(row = 1, column = 4).value = 'Root trace projection radius'
        sheet.cell(row = 1, column = 5).value = 'Root number in total'
    
    sheet = wb.active

    for row in zip(index, edge_length, edge_angle, projection_radius):
        sheet.append(row)
    
    
    sheet.cell(row = 2, column = 5).value = numer_total
        
    #save the csv file
    wb.save(trait_file)
    
    if os.path.exists(trait_file):
        
        print("Trait result was saved!\n")
    
    import openpyxl
    import csv

    wb = load_workbook(trait_file)
    sh = wb.get_active_sheet()
    
    #with open(trait_file_csv, 'wb') as f:
    with open(trait_file_csv, 'w', newline = "") as f:
        c = csv.writer(f)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])
    
    
    ##################################################################
    #End of writing measured parameters as excel file 
    
    
