"""
Version: 1.5

Summary: Analyze and visualzie tracked traces

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE:

python3 track_load.py -p /home/suxingliu/Ptvpy_test/ -v True


argument:
("-p", "--path", required = True, help="path to trace file")
("-v", "--visualize", required = False, default = False, type = bool, help = "Visualize result or not")

default file format: *.csv 

"""

# Import python libraries
import numpy as np
from numpy import arctan2, sqrt
#import numexpr as ne

#import matplotlib as mpl
#import matplotlib.cm as cm
#import matplotlib.pyplot as plt

import glob
import fnmatch
import os, os.path
import math

import argparse

from openpyxl import load_workbook
from openpyxl import Workbook
        
from scipy.spatial import distance

from mayavi import mlab

import itertools

#import warnings
#warnings.filterwarnings("ignore", category=FutureWarning)
#warnings.filterwarnings("ignore")

from tabulate import tabulate
import pandas as pd



from skimage.morphology import skeletonize
import sknw
import matplotlib.pyplot as plt

from mpl_toolkits.mplot3d import axes3d, Axes3D 

from skimage.morphology import skeletonize_3d
from network_3d import skel2graph, plot_graph
from networkx import nx

import dask
import dask.array as da

def mkdir(path):
    """Create result folder"""
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False

#colormap mapping
def get_cmap(n, name='hsv'):
    """get the color mapping"""
    
    #Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    #RGB color; the keyword argument name must be a standard mpl colormap name
    return plt.cm.get_cmap(name,n+1)
    
  
'''
def connect_trace(trace_array, trace_index, trace_number):
    
    if args["visualize"]:
        
        f = mlab.figure(bgcolor = (1,1,1), fgcolor = (0.5, 0.5, 0.5), size = (600,400))

        cmap = get_cmap(trace_number)
    
    
    #initialize parameters
    index_rec = []
    length_rec = []
    angle_rec = []
    diameter_rec = []
    projection_radius = []
    
    
    for idx, index_value in enumerate(trace_index):

        #print(idx, index_value)
        
        X = trace_array[np.where(trace_array[:,0] == index_value)][:,1]
        Y = trace_array[np.where(trace_array[:,0] == index_value)][:,2]
        Z = trace_array[np.where(trace_array[:,0] == index_value)][:,3]/1
        
        #traits measurement
        ##################################################################
        radius_mean = np.mean(trace_array[np.where(trace_array[:,0] == index_value)][:,4])
        
        scalars = trace_array[np.where(trace_array[:,0] == index_value)][:,4]

        #compute line length
        coords = np.stack(( X, Y, Z ), axis = 1)
        
        #line_length = pathlength(X, Y, Z)
        
        line_length = points_seg_length(coords)
        
        #(r, theta, phi) = points_angle(X, Y, Z, line_length)
        
        (r, azimuth, elevation) = cart2sph(X, Y, Z)
        
        if azimuth > 90:
            angle = 180 - azimuth
        elif azimuth < 0:
            angle = 90 + azimuth
        else:
            angle = azimuth

        print("Trace {0} properities:".format(idx))
        
        print("Number of points:{}, Length:{:.2f}, Angle:{:.2f} \n".format(len(X), line_length, angle))
        
        #print("Angle:{0} {1} {2}\n".format(r, theta, phi))
        
        
            
        # record all parameters
        index_rec.append(idx)
        length_rec.append(line_length)
        angle_rec.append(angle)
        diameter_rec.append(radius_mean)
        projection_radius.append(r)
        
        
        ####################################################################Visualiztion
        if args["visualize"]:
        
            #generate random different colors for different traces
            color_rgb = cmap(idx)[:len(cmap(idx))-1]
                    
            #Draw 3d points and lines 
            pts = mlab.points3d(X, Y, Z, scalars, color = color_rgb, mode = 'point')
            pts.actor.property.set(point_size = 5.5)
            
            #pts = mlab.plot3d(X, Y, Z, color = color_rgb, opacity = 0.5, representation = 'wireframe', transparent = True, tube_radius = radius)
            
            #pts = mlab.plot3d(X, Y, Z, color = color_rgb, opacity = 0.5, representation = 'points', transparent = True, tube_radius = radius)
            
            pts = mlab.plot3d(X, Y, Z, color = color_rgb, opacity = 0.3, representation = 'surface', transparent = True, tube_radius = radius_mean)
        
        
        ####################################################################Visualiztion
    
    if args["visualize"]:
        #show model
        mlab.show()
    
    return index_rec, length_rec, angle_rec, diameter_rec, projection_radius
'''


# compute the path length along the trace
def pathlength(x,y,z):

    n = len(x)
    
    lv = [sqrt((x[i]-x[i-1])**2 + (y[i]-y[i-1])**2 + (z[i]-z[i-1])**2) for i in range(n)]
    
    return sum(lv)

# compute distnace between consective point sets
def points_seg_length(coords):
    
    d = np.diff(coords, axis=0)
    
    segdists = np.sqrt((d ** 2).sum( axis = 1))
    
    # calculate length of line
    #l = np.sqrt( np.diff(X)**2 + np.diff(Y)**2 + np.diff(Z)**2 )
    
    return sum(segdists)


# compute angle between two 3D points
def points_angle(x, y, z, line_length):
    
    theta_offest = np.zeros(4)
    r_offest = np.zeros(4)
        
    #calculate angles
    for offest in range(0,4):
        
        interval = int(((offest+1)*0.25)*line_length)  
        
        if interval >= len(x):
            interval = len(x)-1
        
        cx = x[interval] - x[0]
        cy = y[interval] - y[0]  
        cz = z[interval] - z[0]
        
        (r,theta,phi) = asSpherical(cx, cy, cz)
        
        if theta > 90:
            theta = 180 -theta

        theta_offest[offest] = theta

        r_offest[offest] = r

    return r_offest[2], theta_offest[2], phi



#coordinates transformation from cartesian coords to sphere coord system
def cart2sph(x, y, z):
    hxy = np.hypot(x, y)
    r = np.hypot(hxy, z)
    elevation = np.arctan2(z, hxy)*180/math.pi
    azimuth = np.arctan2(y, x)*180/math.pi
    return r[2], azimuth[2], elevation[2]

#coordinates transformation from cartesian coords to sphere coord system
def appendSpherical_np(xyz):
    
    ptsnew = np.hstack((xyz, np.zeros(xyz.shape)))
    xy = xyz[:,0]**2 + xyz[:,1]**2
    ptsnew[:,3] = np.sqrt(xy + xyz[:,2]**2)
    ptsnew[:,4] = np.arctan2(np.sqrt(xy), xyz[:,2]) # for elevation angle defined from Z-axis down
    #ptsnew[:,4] = np.arctan2(xyz[:,2], np.sqrt(xy)) # for elevation angle defined from XY-plane up
    ptsnew[:,5] = np.arctan2(xyz[:,1], xyz[:,0])
    return ptsnew[:,3],ptsnew[:,4],ptsnew[:,5]


# visualize the trace in 2D and apply color coding for each trace
def visualize_trace_mayavi(trace_array, trace_index, trace_number):
    
    if args["visualize"]:
        
        f = mlab.figure(bgcolor = (1,1,1), fgcolor = (0.5, 0.5, 0.5), size = (600,400))

        cmap = get_cmap(trace_number)
    
    
    '''
    X = [0,1,2]
    Y = [0,1,2]
    Z = [0,1,2]
    
    #stack to array
    coords = np.stack(( X, Y, Z ), axis = 1)

    #compute line length
    line_length = points_seg_length(coords)

    #compute angle
    (r, theta, phi) = cart2sph(X, Y, Z)
    
    print("line_length, r, theta, phi")
    print(line_length, r, theta, phi)
    '''
    
    
    
    #initialize parameters
    index_rec = []
    length_rec = []
    angle_rec = []
    diameter_rec = []
    projection_radius = []
    
    image_chunk = np.zeros((413, 411, 30))
    
    new_mask = np.zeros_like(image_chunk)

   
    
    for idx, index_value in enumerate(trace_index):

        #print(idx, index_value)
        
        X = trace_array[np.where(trace_array[:,0] == index_value)][:,1]
        Y = trace_array[np.where(trace_array[:,0] == index_value)][:,2]
        Z = trace_array[np.where(trace_array[:,0] == index_value)][:,3]/1
        
        image_chunk[X.astype(int), Y.astype(int), Z.astype(int)] = 1
        
        
        #traits measurement
        ##################################################################
        radius_mean = np.mean(trace_array[np.where(trace_array[:,0] == index_value)][:,4])
        
        scalars = trace_array[np.where(trace_array[:,0] == index_value)][:,4]

        #compute line length
        coords = np.stack(( X, Y, Z ), axis = 1)
        
        #line_length = pathlength(X, Y, Z)
        
         #compute line angle
        line_length = points_seg_length(coords)
        
        (r, azimuth, elevation) = cart2sph(X, Y, Z)
        
        if azimuth > 90:
            angle = 180 - azimuth
        elif azimuth < 0:
            angle = 90 + azimuth
        else:
            angle = azimuth

        print("Trace {0} properities:".format(idx))
        
        print("Number of points:{}, Length:{:.2f}, Angle:{:.2f} \n".format(len(X), line_length, angle))
        
        #print("Angle:{0} {1} {2}\n".format(r, theta, phi))
        
        
            
        # record all parameters
        index_rec.append(idx)
        length_rec.append(line_length)
        angle_rec.append(angle)
        diameter_rec.append(radius_mean)
        projection_radius.append(r)
        
        
        ####################################################################Visualiztion
        if args["visualize"]:
        
            #generate random different colors for different traces
            color_rgb = cmap(idx)[:len(cmap(idx))-1]
                    
            #Draw 3d points and lines 
            pts = mlab.points3d(X, Y, Z, scalars, color = color_rgb, mode = 'point')
            pts.actor.property.set(point_size = 5.5)
            
            #pts = mlab.plot3d(X, Y, Z, color = color_rgb, opacity = 0.5, representation = 'wireframe', transparent = True, tube_radius = radius)
            
            #pts = mlab.plot3d(X, Y, Z, color = color_rgb, opacity = 0.5, representation = 'points', transparent = True, tube_radius = radius)
            
            pts = mlab.plot3d(X, Y, Z, color = color_rgb, opacity = 0.3, representation = 'surface', transparent = True, tube_radius = radius_mean)
        
        
        ####################################################################Visualiztion
    
    if args["visualize"]:
        #show model
        mlab.show()
    
    return index_rec, length_rec, angle_rec, diameter_rec, projection_radius, image_chunk
    
    

#visualize graph using mayavi
def mayavi_visualize(graph,image_chunk):
    
    mlab.figure(1, size = (500, 500))
    
    cmap = get_cmap(len(graph.edges()))
    
    #for (s,e) in graph.edges():
        
    for idx, (s,e) in enumerate(graph.edges()):
        
         #generate different colors for different traces
        color_rgb = cmap(idx)[:len(cmap(idx))-1]
        
        pst = graph[s][e]['pts']

        mlab.plot3d(pst[:,0], pst[:,1], pst[:,2], tube_radius = 0.025, color = color_rgb)
    
    nodes = graph.nodes()
    
    ps = np.array([nodes[i]['o'] for i in nodes])
    
    mlab.points3d(ps[:,0], ps[:,1], ps[:,2], color=(1.0, 0.0, 0.0), scale_factor = 2.75)
    
    mlab.contour3d(image_chunk.astype(np.float),contours=[.5],opacity=.5,color=(1,1,1))
    
    mlab.show() 




if __name__ == '__main__':
    
    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True, help = "path to trace file")
    ap.add_argument("-v", "--visualize", required = False, default = False, type = bool, help = "Visualize result or not")
    #ap.add_argument("-dt", "--dis_tracking", required = False, default = '50.5', type = float, help = "dis_tracking")
    #ap.add_argument("-ma", "--min_angle", required = False, default = '0.1', type = float, help = "min_angle")
    #ap.add_argument("-dr", "--dist_ratio", required = False, default = '4.8', type = float, help = "dist_ratio")
    args = vars(ap.parse_args())

    #extract file list in specified path
    filetype = '*.csv' 
    file_path = args["path"]
    
    #accquire file list
    file_list = sorted(fnmatch.filter(os.listdir(args["path"]), filetype))
      
    #file_list.sort(key=lambda f: int(filter(str.isdigit, f)))
    
    print(file_list)
    
    global trace_rec, dis_tracking, min_angle, dist_ratio
    
    trace_rec = []
    
    #define min distance tracking threshold
    #dis_tracking = args["dis_tracking"]
    #min_angle = args["min_angle"]
    #dist_ratio = args["dist_ratio"]

    
    # make the folder to store the results
    parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    mkpath = parent_path + '/' + str('analysis_result')
    mkdir(mkpath)
    save_path_result = mkpath + '/'
  
    
    #loop to all tracked trace files
    for file_idx, fname in enumerate(file_list):

        df = pd.read_csv(fname)
        
        trace_pd = df[['particle', 'x', 'y', 'frame', 'size', 'mass', 'raw_mass']]
        
        trace_index = trace_pd["particle"].unique()
        
        trace_number = len(trace_index)

        
        
    # convert pd foramt to nump array
    trace_array = trace_pd.to_numpy()
    
    # trait computation and visualzition
    (index_rec, length_rec, angle_rec, diameter_rec, projection_radius, image_chunk) = visualize_trace_mayavi(trace_array, trace_index, trace_number)
    
    print(image_chunk.shape)
    
    skel = skeletonize(image_chunk)
    
    skel = skel.astype(np.bool) #data needs to be bool
    
    G = skel2graph(skel) #create graph
    
    #print(G.size())
    
    
    #compute edge properities 
    numer_total = G.size()
    
    #plot the graph, use the z component to colorcode both the edges and the nodes, scale nodes according to their degree
    #plot_graph(G,node_color_keyword='z',edge_color_keyword='z',scale_node_keyword='degree')
    
    edge_node_n1_select, edge_node_n2_select, angle_select, length_select, projection_select = plot_graph(G, node_color_keyword = 'x', edge_color_keyword = 'x')
    
    #tube_surf, pts, edge_node_n1_select, edge_node_n2_select, angle_select, length_select, projection_select = plot_graph(G, node_color_keyword = 'z', edge_color_keyword = 'z')
    
    #print("edge_node_unique: {0}\n".format(edge_node_unique))

    index = []
    edgecount = len(edge_node_n1_select)
    
    for i in range(edgecount):
        
        n1 = edge_node_n1_select[i]
        n2 = edge_node_n2_select[i]
        
        try:
            print("Properities of root index {0}:".format(i))
            print("node1 = {0}, node2 = {1} ".format(n1, n2))
            print("angle = {0}, length = {1} ".format(angle_select[i], length_select[i]))
            print("projection_radius = {0}\n ".format(projection_select[i]))
            
            index.append(i+1)
            
        except IndexError:
            pass
        
    
    
                
    #show the binary data
    #mlab.contour3d(image_chunk.astype(np.float),contours=[.5],opacity=.5,color=(1,1,1))
    
    mlab.show()

    
    
    #output total number of traces detected
    print("Summary: {0} unique root trajectories were detected...\n".format(trace_number))
    
    #output in command window in a sum table
    trait_sum = []
    
    for row in zip(index_rec, length_rec, angle_rec, diameter_rec, projection_radius):
       trait_sum.append(row)

    table = tabulate(trait_sum, headers = ['Root trace index', 'Length', 'Angle', 'Diameter', 'Projection radius' ], tablefmt = 'orgtbl')

    print(table + "\n")
    '''
    ##################################################################
    #Start of writing measured parameters as excel file 

    #parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    #base_folder = os.path.basename(file_path[:-1])
    #trait_file = (parent_path + '/' + base_folder + 'root_trace_measure' + '.xlsx')
    #trait_file_csv = (parent_path + '/' + base_folder + 'root_trace_measure' + '.csv')
    
    trait_file = (save_path_result + 'root_trace_measure' + '.xlsx')
    
    trait_file_csv = (save_path_result + 'root_trace_measure' + '.csv')
    
    if os.path.exists(trait_file):
        # update values
        #Open an xlsx for reading
        wb = load_workbook(trait_file, read_only = False)
        sheet = wb.active

        os.remove(trait_file)
        
    else:
        # Keep presents
        wb = Workbook()
        sheet = wb.active
        
        sheet.cell(row = 1, column = 1).value = 'Root trace index'
        sheet.cell(row = 1, column = 2).value = 'Root trace length'
        sheet.cell(row = 1, column = 3).value = 'Root trace angle'
        sheet.cell(row = 1, column = 4).value = 'Root trace diameter'
        sheet.cell(row = 1, column = 5).value = 'Root trace projection radius'
        sheet.cell(row = 1, column = 6).value = 'Root number in total'
    
    sheet = wb.active

    for row in zip(index_rec, length_rec, angle_rec, diameter_rec, projection_radius):
        sheet.append(row)
    
    sheet.cell(row = 2, column = 6).value = trace_number
    
    #save the csv file
    wb.save(trait_file)
    
    if os.path.exists(trait_file):
        
        print("Trait result was saved in: " + trait_file + "\n")
    
    import openpyxl
    import csv

    wb = load_workbook(trait_file)
    sh = wb.get_active_sheet()
    
    #with open(trait_file_csv, 'wb') as f:
    with open(trait_file_csv, 'w', newline = "") as f:
        c = csv.writer(f)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])
    
    
    ##################################################################
    #End of writing measured parameters as excel file 
    
    '''
   






